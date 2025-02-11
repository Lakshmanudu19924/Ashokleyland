import streamlit as st
import pyrebase
import firebase_admin
from firebase_admin import credentials, auth
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
from io import BytesIO
unique_key = str(uuid.uuid4())
import re


firebaseConfig = {
  "apiKey": "AIzaSyAcQpRtTUZ_LH9MqKuV0IJ51zqD_tAU410",
  "authDomain": "ashok-88fd9.firebaseapp.com",
  "projectId": "ashok-88fd9",
  "storageBucket": "ashok-88fd9.firebasestorage.app",
  "messagingSenderId": "249675153769",
  "appId": "1:249675153769:web:c59ed35637fcc4b9dc7df8",
  "measurementId": "G-PQFTZ53R4M",
  "databaseURL": "https://ashok-88fd9-default-rtdb.firebaseio.com"
}

firebase = pyrebase.initialize_app(firebaseConfig)
auth_client = firebase.auth()

if not firebase_admin._apps:
    cred = credentials.Certificate("ashok-88fd9-firebase-adminsdk-fbsvc-a87adc24e7.json")  # Replace with your Firebase Admin SDK JSON file
    firebase_admin.initialize_app(cred)



def login():
    st.title("Login")

    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        try:
            user = auth_client.sign_in_with_email_and_password(email, password)
            st.success("Successfully logged in!")
            st.session_state["user"] = user
        except Exception as e:
            st.error(f"Error: {e}")


def register():
    st.title("Register")

    email = st.text_input("Email")
    password = st.text_input("Password", type="password")
    confirm_password = st.text_input("Confirm Password", type="password")

    if st.button("Register"):
        # Check if the email domain matches
        if not email.endswith("@ashokleyland.com"):
            st.error("Registration is restricted to @ashokleyland.com email addresses.")
        elif password != confirm_password:
            st.error("Passwords do not match!")
        else:
            try:
                auth_client.create_user_with_email_and_password(email, password)
                st.success("Successfully registered! Please log in.")
            except Exception as e:
                st.error(f"Error: {e}")


def logout():
    if "user" in st.session_state:
        del st.session_state["user"]
        st.success("Logged out successfully!")
        


def app_functionality():
    # Ensure user is logged in
    if "user" not in st.session_state:
        st.warning("Please log in to access the app.")
        st.stop()

    # Initialize session state variables
    if "selected_option" not in st.session_state:
        st.session_state.selected_option = None

    st.sidebar.title("Choose a Functionality")

    # Define available options
    options = {
        "Mapped set available without considering alternates": map_wout_alt,
        "Mapped set available considering alternates": map_w_alt,
        
        "MPS Plan - 2 Weeks with Alternates": two_week_w_al,
        "MPS Plan - 2 Weeks without Alternates": two_week_wo_al,
     
        "MPS Plan - 4 Weeks with Alternates": four_week_with_alter,
        "MPS Plan - 4 Weeks without Alternates": four_week_without_alter,
        "Part Calculation": process_part_matrix_master,
        "Priority Sheet": Priority_Analysis_P_NO_with_WIP_Description_and_SUB1_Mapping,
        "Month GB Req After OS": Month,
        "GB Req for Balance Month": Gbreq
    }

    # Function to update selection
    def update_selection(selection):
        st.session_state.selected_option = selection

    # Sidebar expander for Matched Set
    
    # Sidebar expander for Additional Calculations
    with st.sidebar.expander("Additional Calculations"):
        if st.button("Month GB Req After OS", key="month_gb_req"):
            update_selection("Month GB Req After OS")
        if st.button("GB Req for Balance Month", key="gb_req_bal_month"):
            update_selection("GB Req for Balance Month")
    
        if st.button("Made Here Part Calculation", key="part_calc"):
            update_selection("Part Calculation")
        if st.button("Priority Sheet", key="priority_sheet"):
            update_selection("Priority Sheet")
       
    with st.sidebar.expander("Matched Set"):
        st.markdown("**Against Tentative Plan**")
        if st.button("Without Alternates", key="without_alt"):
            update_selection("Mapped set available without considering alternates")
        if st.button("With Alternates", key="with_alt"):
            update_selection("Mapped set available considering alternates")

        st.markdown("**Against MPS - 2 Weeks**")
        
        if st.button("2-Week Plan with Alternates", key="2_week_alt"):
            update_selection("MPS Plan - 2 Weeks with Alternates")
        if st.button("2-Week Plan without Alternates", key="2_week_wout_alt"):
            update_selection("MPS Plan - 2 Weeks without Alternates")

        st.markdown("**Against MPS - 4 Weeks**")
        
        if st.button("4-Week Plan with Alternates", key="4_week_alt"):
            update_selection("MPS Plan - 4 Weeks with Alternates")
        if st.button("4-Week Plan without Alternates", key="4_week_wout_alt"):
            update_selection("MPS Plan - 4 Weeks without Alternates")


    # **Render the selected functionality**
    selected = st.session_state.selected_option
    if selected:
        st.write(f"### Running: {selected}")
        if selected in options and options[selected]:  
            options[selected]()  # Call the corresponding function
        else:
            st.warning("Functionality not yet implemented.")
            
            
def two_week_w_al():
    # Title of the app
    st.title('2-week-with-alternative')

    # File uploader widget to upload the Excel file
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file is not None:
        # Load the Excel file
        try:
            # Load all sheet names and match case-insensitively by converting them to lowercase
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names
            
            # Convert sheet names to lowercase for case-insensitive comparison
            target_sheet_name_1 = "GB Requirement for Bal Month".lower()
            target_sheet_name_2 = "Part Raw Data".lower()
            target_sheet_name_3 = "Made Here Parts Calc".lower()
            target_sheet_name_4 = "Alternate Part Master".lower()
            
            matching_sheets_1 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_1]
            matching_sheets_2 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_2]
            matching_sheets_3 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_3]
            matching_sheets_4 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_4]
            
            if matching_sheets_1 and matching_sheets_2 and matching_sheets_3 and matching_sheets_4:
                # Read the matched sheets
                df_1 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_1[0])
                df_2 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_2[0])
                df_3 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_3[0])
                df_4 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_4[0])

                # Clean column names: remove leading/trailing spaces and make them case-insensitive
                df_1.columns = df_1.columns.str.strip().str.lower()
                df_2.columns = df_2.columns.str.strip().str.lower()
                df_3.columns = df_3.columns.str.strip().str.lower()
                df_4.columns = df_4.columns.str.strip().str.lower()

                # Check if 'spe' column exists in both sheets
                if 'spe' in df_1.columns and 'spe' in df_2.columns:
                    # Merge the two dataframes based on 'spe'
                    merged_data = pd.merge(df_1[['spe', 'w2 rev']], df_2, on='spe', how='inner')
                    
                    # List of the required columns from the Part Raw Data sheet
                    required_columns = [
                        '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                        'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                        'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                        'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                    ]

                    # Check if all required columns are present
                    missing_columns = [col for col in required_columns if col not in merged_data.columns]
                    if missing_columns:
                        st.error(f"Missing columns in Part Raw Data: {', '.join(missing_columns)}")
                    else:
                        # Select only the relevant columns and add Serial Number starting from 1
                        filtered_data = merged_data[['spe', 'w2 rev'] + required_columns]
                        
                        # Replace None or NaN with 0 in the filtered data
                        filtered_data = filtered_data.fillna(0)

                        # Add Serial Number starting from 1
                        filtered_data.insert(0, 'Serial Number', range(1, len(filtered_data) + 1))

                        # --- New functionality for "Made Here Parts Calc" sheet ---

                        # Extract P.NO and CURRENT MH columns from "Made Here Parts Calc"
                        if 'p.no' in df_3.columns and 'current mh' in df_3.columns:
                            made_here_parts_calc_df = df_3[['p.no', 'current mh']]

                            # Track remaining stock using P.NO
                            remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                            # Function to calculate CURRENT MH and REMAINING MH row-wise for each part and column
                            def calculate_row_current_and_remaining(row, column):
                                part_id = row[column]
                                if part_id in remaining_stock:
                                    available_mh = remaining_stock[part_id]
                                    current_mh = min(row["w2 rev"], available_mh)  # Choose the min between W2 REV and available MH
                                    remaining_stock[part_id] -= current_mh  # Update remaining stock
                                    return current_mh
                                return 0

                            # Columns to process (the ones that start with 'F' like '1st on ms', '2nd on ms', etc.)
                            columns_to_process = [
                                '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                                'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                                'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                                'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                            ]

                            # Calculate CURRENT MH and REMAINING MH for all components
                            for col in columns_to_process:
                                # Calculate CURRENT MH for each part (using the `calculate_row_current_and_remaining` function)
                                filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                    lambda row: calculate_row_current_and_remaining(row, col), axis=1
                                )

                                # Calculate REMAINING MH as W2 REV - CURRENT MH
                                filtered_data[f"REMAINING MH ({col})"] = (
                                    filtered_data["w2 rev"] - filtered_data[f"CURRENT MH ({col})"]
                                )

                            # --- New functionality for "Alternate Part Master" sheet ---

                            if 'p.no' in df_4.columns and 'sub1' in df_4.columns and 'sub2' in df_4.columns:
                                alternate_part_master_df = df_4[['p.no', 'sub1', 'sub2']]

                                # Step 3: Map P.NO to SUB1 and SUB2 for each column
                                alternate_part_dict = alternate_part_master_df.set_index("p.no")[["sub1", "sub2"]].to_dict("index")

                                for col in columns_to_process:
                                    filtered_data[f"SUB1 ({col})"] = filtered_data[col].map(
                                        lambda x: alternate_part_dict[x]["sub1"] if x in alternate_part_dict else 0
                                    )
                                    filtered_data[f"SUB2 ({col})"] = filtered_data[col].map(
                                        lambda x: alternate_part_dict[x]["sub2"] if x in alternate_part_dict else 0
                                    )

                                # Step 4: Calculate CURRENT MH for each column
                                remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                                def calculate_row_current(row, column):
                                    key = row[column]
                                    if key in remaining_stock:
                                        available_mh = remaining_stock[key]
                                        used_mh = min(row["w2 rev"], available_mh)
                                        remaining_stock[key] -= used_mh
                                        return used_mh
                                    return 0

                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                        lambda row: calculate_row_current(row, col), axis=1
                                    )

                                # Step 5: Calculate CURRENT MH for SUB1
                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH (SUB1 {col})"] = filtered_data.apply(
                                        lambda row: max(row["w2 rev"] - row[f"CURRENT MH ({col})"], 0), axis=1
                                    )

                                # Step 6: Calculate CURRENT MH for SUB2
                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH (SUB2 {col})"] = filtered_data.apply(
                                        lambda row: max(row[f"CURRENT MH ({col})"] + row[f"CURRENT MH (SUB1 {col})"] - row["w2 rev"], 0), axis=1
                                    )

                            else:
                                st.error('Columns "P.NO", "SUB1" or "SUB2" not found in the "Alternate Part Master" sheet.')

                            # --- Step 7: Calculate the minimum CURRENT MH for each row, excluding zero-value columns ---
                            def calculate_min_current_mh(row):
                                non_zero_values = [
                                    row[f"CURRENT MH ({col})"] for col in columns_to_process
                                    if row[col] != 0
                                ]
                                return min(non_zero_values) if non_zero_values else 0

                            filtered_data["MINIMUM CURRENT MH"] = filtered_data.apply(
                                calculate_min_current_mh, axis=1
                            )

                            # Replace any remaining NaN or None values with 0
                            filtered_data = filtered_data.fillna(0)

                            # Re-arranging the columns as per the given order
                            final_columns = [
                                'spe', 'w2 rev', 'MINIMUM CURRENT MH', 
                                '1st on ms', 'CURRENT MH (1st on ms)', 'SUB1 (1st on ms)', 'CURRENT MH (SUB1 1st on ms)', 'SUB2 (1st on ms)', 'CURRENT MH (SUB2 1st on ms)',
                                '2nd on ms', 'CURRENT MH (2nd on ms)', 'SUB1 (2nd on ms)', 'CURRENT MH (SUB1 2nd on ms)', 'SUB2 (2nd on ms)', 'CURRENT MH (SUB2 2nd on ms)',
                                '3rd on ms', 'CURRENT MH (3rd on ms)', 'SUB1 (3rd on ms)', 'CURRENT MH (SUB1 3rd on ms)', 'SUB2 (3rd on ms)', 'CURRENT MH (SUB2 3rd on ms)',
                                '4th on ms', 'CURRENT MH (4th on ms)', 'SUB1 (4th on ms)', 'CURRENT MH (SUB1 4th on ms)', 'SUB2 (4th on ms)', 'CURRENT MH (SUB2 4th on ms)',
                                '5th on ms', 'CURRENT MH (5th on ms)', 'SUB1 (5th on ms)', 'CURRENT MH (SUB1 5th on ms)', 'SUB2 (5th on ms)', 'CURRENT MH (SUB2 5th on ms)',
                                'rev on ms', 'CURRENT MH (rev on ms)', 'SUB1 (rev on ms)', 'CURRENT MH (SUB1 rev on ms)', 'SUB2 (rev on ms)', 'CURRENT MH (SUB2 rev on ms)',
                                'cm on ls', 'CURRENT MH (cm on ls)', 'SUB1 (cm on ls)', 'CURRENT MH (SUB1 cm on ls)', 'SUB2 (cm on ls)', 'CURRENT MH (SUB2 cm on ls)',
                                'rev idler', 'CURRENT MH (rev idler)', 'SUB1 (rev idler)', 'CURRENT MH (SUB1 rev idler)', 'SUB2 (rev idler)', 'CURRENT MH (SUB2 rev idler)',
                                '3rd on ls', 'CURRENT MH (3rd on ls)', 'SUB1 (3rd on ls)', 'CURRENT MH (SUB1 3rd on ls)', 'SUB2 (3rd on ls)', 'CURRENT MH (SUB2 3rd on ls)',
                                '4th on ls', 'CURRENT MH (4th on ls)', 'SUB1 (4th on ls)', 'CURRENT MH (SUB1 4th on ls)', 'SUB2 (4th on ls)', 'CURRENT MH (SUB2 4th on ls)',
                                '5th on ls', 'CURRENT MH (5th on ls)', 'SUB1 (5th on ls)', 'CURRENT MH (SUB1 5th on ls)', 'SUB2 (5th on ls)', 'CURRENT MH (SUB2 5th on ls)',
                                'input shaft', 'CURRENT MH (input shaft)', 'SUB1 (input shaft)', 'CURRENT MH (SUB1 input shaft)', 'SUB2 (input shaft)', 'CURRENT MH (SUB2 input shaft)',
                                'main shaft', 'CURRENT MH (main shaft)', 'SUB1 (main shaft)', 'CURRENT MH (SUB1 main shaft)', 'SUB2 (main shaft)', 'CURRENT MH (SUB2 main shaft)',
                                'lay shaft', 'CURRENT MH (lay shaft)', 'SUB1 (lay shaft)', 'CURRENT MH (SUB1 lay shaft)', 'SUB2 (lay shaft)', 'CURRENT MH (SUB2 lay shaft)',
                                'hub 1/ 2', 'CURRENT MH (hub 1/ 2)', 'SUB1 (hub 1/ 2)', 'CURRENT MH (SUB1 hub 1/ 2)', 'SUB2 (hub 1/ 2)', 'CURRENT MH (SUB2 hub 1/ 2)',
                                'hub 3/4', 'CURRENT MH (hub 3/4)', 'SUB1 (hub 3/4)', 'CURRENT MH (SUB1 hub 3/4)', 'SUB2 (hub 3/4)', 'CURRENT MH (SUB2 hub 3/4)',
                                'hub 5/6', 'CURRENT MH (hub 5/6)', 'SUB1 (hub 5/6)', 'CURRENT MH (SUB1 hub 5/6)', 'SUB2 (hub 5/6)', 'CURRENT MH (SUB2 hub 5/6)',
                                'fdr', 'CURRENT MH (fdr)', 'SUB1 (fdr)', 'CURRENT MH (SUB1 fdr)', 'SUB2 (fdr)', 'CURRENT MH (SUB2 fdr)',
                                'sleeve 1/ 2', 'CURRENT MH (sleeve 1/ 2)', 'SUB1 (sleeve 1/ 2)', 'CURRENT MH (SUB1 sleeve 1/ 2)', 'SUB2 (sleeve 1/ 2)', 'CURRENT MH (SUB2 sleeve 1/ 2)',
                                'sleeve 3/4', 'CURRENT MH (sleeve 3/4)', 'SUB1 (sleeve 3/4)', 'CURRENT MH (SUB1 sleeve 3/4)', 'SUB2 (sleeve 3/4)', 'CURRENT MH (SUB2 sleeve 3/4)',
                                'sleeve 5/6', 'CURRENT MH (sleeve 5/6)', 'SUB1 (sleeve 5/6)', 'CURRENT MH (SUB1 sleeve 5/6)', 'SUB2 (sleeve 5/6)', 'CURRENT MH (SUB2 sleeve 5/6)',
                                'cone 1/2', 'CURRENT MH (cone 1/2)', 'SUB1 (cone 1/2)', 'CURRENT MH (SUB1 cone 1/2)', 'SUB2 (cone 1/2)', 'CURRENT MH (SUB2 cone 1/2)',
                                'cone 3/4', 'CURRENT MH (cone 3/4)', 'SUB1 (cone 3/4)', 'CURRENT MH (SUB1 cone 3/4)', 'SUB2 (cone 3/4)', 'CURRENT MH (SUB2 cone 3/4)',
                                'cone 5/6', 'CURRENT MH (cone 5/6)', 'SUB1 (cone 5/6)', 'CURRENT MH (SUB1 cone 5/6)', 'SUB2 (cone 5/6)', 'CURRENT MH (SUB2 cone 5/6)',
                                'cone 3', 'CURRENT MH (cone 3)', 'SUB1 (cone 3)', 'CURRENT MH (SUB1 cone 3)', 'SUB2 (cone 3)', 'CURRENT MH (SUB2 cone 3)',
                                'cone 4', 'CURRENT MH (cone 4)', 'SUB1 (cone 4)', 'CURRENT MH (SUB1 cone 4)', 'SUB2 (cone 4)', 'CURRENT MH (SUB2 cone 4)'
                            ]

                            # Reorder columns based on final_columns order
                            filtered_2_w_al = filtered_data[final_columns]
                            

                            # Display the final DataFrame
                            st.dataframe(filtered_2_w_al)
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Processed Data"

                                # Write DataFrame to the worksheet
                            for row in dataframe_to_rows(filtered_2_w_al, index=False, header=True):
                                ws.append(row)

                                # Save the workbook to the BytesIO object
                            wb.save(output)
                            processed_file = output.getvalue()

                            st.download_button(
                                label="Download Processed Excel",
                                data=processed_file,
                                file_name="2_week_with_alternative.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                        else:
                            st.error('Error: Missing "P.NO", "SUB1", or "SUB2" columns in the Alternate Part Master sheet.')

        except Exception as e:
            st.error(f"An error occurred: {e}")

def two_week_wo_al():
    # Title of the app
    st.title('2-week-without-alternative')

    # File uploader widget to upload the Excel file
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file is not None:
        # Load the Excel file
        try:
            # Load all sheet names and match case-insensitively by converting them to lowercase
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names

            # Convert sheet names to lowercase for case-insensitive comparison
            target_sheet_name_1 = "GB Requirement for Bal Month".lower()
            target_sheet_name_2 = "Part Raw Data".lower()
            target_sheet_name_3 = "Made Here Parts Calc".lower()

            matching_sheets_1 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_1]
            matching_sheets_2 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_2]
            matching_sheets_3 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_3]

            if matching_sheets_1 and matching_sheets_2 and matching_sheets_3:
                # Read the matched sheets
                df_1 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_1[0])
                df_2 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_2[0])
                df_3 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_3[0])

                # Clean column names: remove leading/trailing spaces and make them case-insensitive
                df_1.columns = df_1.columns.str.strip().str.lower()
                df_2.columns = df_2.columns.str.strip().str.lower()
                df_3.columns = df_3.columns.str.strip().str.lower()

                # Check if 'spe' column exists in both sheets
                if 'spe' in df_1.columns and 'spe' in df_2.columns:
                    # Merge the two dataframes based on 'spe'
                    merged_data = pd.merge(df_1[['spe', 'w2 rev']], df_2, on='spe', how='inner')

                    # List of the required columns from the Part Raw Data sheet
                    required_columns = [
                        '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                        'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                        'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                        'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                    ]

                    # Check if all required columns are present
                    missing_columns = [col for col in required_columns if col not in merged_data.columns]
                    if missing_columns:
                        st.error(f"Missing columns in Part Raw Data: {', '.join(missing_columns)}")
                    else:
                        # Select only the relevant columns and add Serial Number starting from 1
                        filtered_data = merged_data[['spe', 'w2 rev'] + required_columns]

                        # Replace None or NaN with 0 in the filtered data
                        filtered_data = filtered_data.fillna(0)

                        # Add Serial Number starting from 1
                        filtered_data.insert(0, 'Serial Number', range(1, len(filtered_data) + 1))

                        # --- New functionality for "Made Here Parts Calc" sheet ---

                        # Extract P.NO and CURRENT MH columns from "Made Here Parts Calc"
                        if 'p.no' in df_3.columns and 'current mh' in df_3.columns:
                            made_here_parts_calc_df = df_3[['p.no', 'current mh']]

                            # Track remaining stock using P.NO
                            remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                            # Function to calculate CURRENT MH and REMAINING MH row-wise for each part and column
                            def calculate_row_current_and_remaining(row, column):
                                part_id = row[column]
                                if part_id in remaining_stock:
                                    available_mh = remaining_stock[part_id]
                                    current_mh = min(row["w2 rev"], available_mh)  # Choose the min between W2 REV and available MH
                                    remaining_stock[part_id] -= current_mh  # Update remaining stock
                                    return current_mh
                                return 0

                            # Columns to process (the ones that start with 'F' like '1st on ms', '2nd on ms', etc.)
                            columns_to_process = [
                                '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                                'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                                'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                                'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                            ]

                            # Calculate CURRENT MH and REMAINING MH for all components
                            for col in columns_to_process:
                                # Calculate CURRENT MH for each part (using the `calculate_row_current_and_remaining` function)
                                filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                    lambda row: calculate_row_current_and_remaining(row, col), axis=1
                                )

                                # Calculate REMAINING MH as W2 REV - CURRENT MH
                                filtered_data[f"REMAINING MH ({col})"] = (
                                    filtered_data["w2 rev"] - filtered_data[f"CURRENT MH ({col})"]
                                )

                            # --- Step 4: Calculate the minimum CURRENT MH for each row ---
                            def calculate_min_current_mh(row):
                                non_zero_values = [
                                    row[f"CURRENT MH ({col})"] for col in columns_to_process
                                    if row[f"CURRENT MH ({col})"] != 0
                                ]
                                return min(non_zero_values) if non_zero_values else 0

                            filtered_data["MINIMUM CURRENT MH"] = filtered_data.apply(
                                calculate_min_current_mh, axis=1
                            )

                            # Re-arranging the columns as per the given order
                            final_columns = [
                                'spe', 'w2 rev','MINIMUM CURRENT MH','1st on ms', 'CURRENT MH (1st on ms)', 'REMAINING MH (1st on ms)',
                                '2nd on ms', 'CURRENT MH (2nd on ms)', 'REMAINING MH (2nd on ms)', '3rd on ms', 'CURRENT MH (3rd on ms)', 'REMAINING MH (3rd on ms)',
                                '4th on ms', 'CURRENT MH (4th on ms)', 'REMAINING MH (4th on ms)', '5th on ms', 'CURRENT MH (5th on ms)', 'REMAINING MH (5th on ms)',
                                'rev on ms', 'CURRENT MH (rev on ms)', 'REMAINING MH (rev on ms)', 'cm on ls', 'CURRENT MH (cm on ls)', 'REMAINING MH (cm on ls)',
                                'rev idler', 'CURRENT MH (rev idler)', 'REMAINING MH (rev idler)', '3rd on ls', 'CURRENT MH (3rd on ls)', 'REMAINING MH (3rd on ls)',
                                '4th on ls', 'CURRENT MH (4th on ls)', 'REMAINING MH (4th on ls)', '5th on ls', 'CURRENT MH (5th on ls)', 'REMAINING MH (5th on ls)',
                                'input shaft', 'CURRENT MH (input shaft)', 'REMAINING MH (input shaft)', 'main shaft', 'CURRENT MH (main shaft)', 'REMAINING MH (main shaft)',
                                'lay shaft', 'CURRENT MH (lay shaft)', 'REMAINING MH (lay shaft)', 'hub 1/ 2', 'CURRENT MH (hub 1/ 2)', 'REMAINING MH (hub 1/ 2)',
                                'hub 3/4', 'CURRENT MH (hub 3/4)', 'REMAINING MH (hub 3/4)', 'hub 5/6', 'CURRENT MH (hub 5/6)', 'REMAINING MH (hub 5/6)',
                                'fdr', 'CURRENT MH (fdr)', 'REMAINING MH (fdr)', 'sleeve 1/ 2', 'CURRENT MH (sleeve 1/ 2)', 'REMAINING MH (sleeve 1/ 2)',
                                'sleeve 3/4', 'CURRENT MH (sleeve 3/4)', 'REMAINING MH (sleeve 3/4)', 'sleeve 5/6', 'CURRENT MH (sleeve 5/6)', 'REMAINING MH (sleeve 5/6)',
                                'cone 1/2', 'CURRENT MH (cone 1/2)', 'REMAINING MH (cone 1/2)', 'cone 3/4', 'CURRENT MH (cone 3/4)', 'REMAINING MH (cone 3/4)',
                                'cone 5/6', 'CURRENT MH (cone 5/6)', 'REMAINING MH (cone 5/6)', 'cone 3', 'CURRENT MH (cone 3)', 'REMAINING MH (cone 3)', 'cone 4', 'CURRENT MH (cone 4)', 'REMAINING MH (cone 4)'
                            ]

                            # Reorder the columns based on the final_columns list
                            filtered_2_wo_al = filtered_data[final_columns]

                            # Display the final mapped data with the CURRENT MH and REMAINING MH columns
                            st.write("Mapped Data from 'GB Requirement for Bal Month', 'Part Raw Data', and 'Made Here Parts Calc':")
                            st.dataframe(filtered_2_wo_al)
                            
                            # Display the final DataFrame
                         
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Processed Data"

                                # Write DataFrame to the worksheet
                            for row in dataframe_to_rows(filtered_2_wo_al, index=False, header=True):
                                ws.append(row)

                                # Save the workbook to the BytesIO object
                            wb.save(output)
                            processed_file = output.getvalue()

                            st.download_button(
                                label="Download Processed Excel",
                                data=processed_file,
                                file_name="2_week_without_alternative.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error('Columns "P.NO" or "CURRENT MH" not found in the "Made Here Parts Calc" sheet.')

                else:
                    st.error('Column "SPE" not found in one or both sheets.')

            else:
                if not matching_sheets_1:
                    st.error('Sheet "GB Requirement for Bal Month" not found in the provided Excel file.')
                if not matching_sheets_2:
                    st.error('Sheet "Part Raw Data" not found in the provided Excel file.')
                if not matching_sheets_3:
                    st.error('Sheet "Made Here Parts Calc" not found in the provided Excel file.')

        except Exception as e:
            st.error(f"Error reading the Excel file: {e}")


def four_week_with_alter():
    st.title('4-week-with-alternative')

# File uploader widget to upload the Excel file
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file is not None:
        # Load the Excel file
        try:
            # Load all sheet names and match case-insensitively by converting them to lowercase
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names
            
            # Convert sheet names to lowercase for case-insensitive comparison
            target_sheet_name_1 = "GB Requirement for Bal Month".lower()
            target_sheet_name_2 = "Part Raw Data".lower()
            target_sheet_name_3 = "Made Here Parts Calc".lower()
            target_sheet_name_4 = "Alternate Part Master".lower()
            
            matching_sheets_1 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_1]
            matching_sheets_2 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_2]
            matching_sheets_3 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_3]
            matching_sheets_4 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_4]
            
            if matching_sheets_1 and matching_sheets_2 and matching_sheets_3 and matching_sheets_4:
                # Read the matched sheets
                df_1 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_1[0])
                df_2 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_2[0])
                df_3 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_3[0])
                df_4 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_4[0])

                # Clean column names: remove leading/trailing spaces and make them case-insensitive
                df_1.columns = df_1.columns.str.strip().str.lower()
                df_2.columns = df_2.columns.str.strip().str.lower()
                df_3.columns = df_3.columns.str.strip().str.lower()
                df_4.columns = df_4.columns.str.strip().str.lower()

                # Check if 'spe' column exists in both sheets
                if 'spe' in df_1.columns and 'spe' in df_2.columns:
                    # Merge the two dataframes based on 'spe'
                    merged_data = pd.merge(df_1[['spe', 'w4 rev']], df_2, on='spe', how='inner')
                    
                    # List of the required columns from the Part Raw Data sheet
                    required_columns = [
                        '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                        'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                        'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                        'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                    ]

                    # Check if all required columns are present
                    missing_columns = [col for col in required_columns if col not in merged_data.columns]
                    if missing_columns:
                        st.error(f"Missing columns in Part Raw Data: {', '.join(missing_columns)}")
                    else:
                        # Select only the relevant columns and add Serial Number starting from 1
                        filtered_data = merged_data[['spe', 'w4 rev'] + required_columns]
                        
                        # Replace None or NaN with 0 in the filtered data
                        filtered_data = filtered_data.fillna(0)

                        # Add Serial Number starting from 1
                        filtered_data.insert(0, 'Serial Number', range(1, len(filtered_data) + 1))

                        # --- New functionality for "Made Here Parts Calc" sheet ---

                        # Extract P.NO and CURRENT MH columns from "Made Here Parts Calc"
                        if 'p.no' in df_3.columns and 'current mh' in df_3.columns:
                            made_here_parts_calc_df = df_3[['p.no', 'current mh']]

                            # Track remaining stock using P.NO
                            remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                            # Function to calculate CURRENT MH and REMAINING MH row-wise for each part and column
                            def calculate_row_current_and_remaining(row, column):
                                part_id = row[column]
                                if part_id in remaining_stock:
                                    available_mh = remaining_stock[part_id]
                                    current_mh = min(row["w4 rev"], available_mh)  # Choose the min between W2 REV and available MH
                                    remaining_stock[part_id] -= current_mh  # Update remaining stock
                                    return current_mh
                                return 0

                            # Columns to process (the ones that start with 'F' like '1st on ms', '2nd on ms', etc.)
                            columns_to_process = [
                                '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                                'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                                'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                                'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                            ]

                            # Calculate CURRENT MH and REMAINING MH for all components
                            for col in columns_to_process:
                                # Calculate CURRENT MH for each part (using the `calculate_row_current_and_remaining` function)
                                filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                    lambda row: calculate_row_current_and_remaining(row, col), axis=1
                                )

                                # Calculate REMAINING MH as W2 REV - CURRENT MH
                                filtered_data[f"REMAINING MH ({col})"] = (
                                    filtered_data["w4 rev"] - filtered_data[f"CURRENT MH ({col})"]
                                )

                            # --- New functionality for "Alternate Part Master" sheet ---

                            if 'p.no' in df_4.columns and 'sub1' in df_4.columns and 'sub2' in df_4.columns:
                                alternate_part_master_df = df_4[['p.no', 'sub1', 'sub2']]

                                # Step 3: Map P.NO to SUB1 and SUB2 for each column
                                alternate_part_dict = alternate_part_master_df.set_index("p.no")[["sub1", "sub2"]].to_dict("index")

                                for col in columns_to_process:
                                    filtered_data[f"SUB1 ({col})"] = filtered_data[col].map(
                                        lambda x: alternate_part_dict[x]["sub1"] if x in alternate_part_dict else 0
                                    )
                                    filtered_data[f"SUB2 ({col})"] = filtered_data[col].map(
                                        lambda x: alternate_part_dict[x]["sub2"] if x in alternate_part_dict else 0
                                    )

                                # Step 4: Calculate CURRENT MH for each column
                                remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                                def calculate_row_current(row, column):
                                    key = row[column]
                                    if key in remaining_stock:
                                        available_mh = remaining_stock[key]
                                        used_mh = min(row["w4 rev"], available_mh)
                                        remaining_stock[key] -= used_mh
                                        return used_mh
                                    return 0

                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                        lambda row: calculate_row_current(row, col), axis=1
                                    )

                                # Step 5: Calculate CURRENT MH for SUB1
                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH (SUB1 {col})"] = filtered_data.apply(
                                        lambda row: max(row["w4 rev"] - row[f"CURRENT MH ({col})"], 0), axis=1
                                    )

                                # Step 6: Calculate CURRENT MH for SUB2
                                for col in columns_to_process:
                                    filtered_data[f"CURRENT MH (SUB2 {col})"] = filtered_data.apply(
                                        lambda row: max(row[f"CURRENT MH ({col})"] + row[f"CURRENT MH (SUB1 {col})"] - row["w4 rev"], 0), axis=1
                                    )

                            else:
                                st.error('Columns "P.NO", "SUB1" or "SUB2" not found in the "Alternate Part Master" sheet.')

                            # --- Step 7: Calculate the minimum CURRENT MH for each row, excluding zero-value columns ---
                            def calculate_min_current_mh(row):
                                non_zero_values = [
                                    row[f"CURRENT MH ({col})"] for col in columns_to_process
                                    if row[col] != 0
                                ]
                                return min(non_zero_values) if non_zero_values else 0

                            filtered_data["MINIMUM CURRENT MH"] = filtered_data.apply(
                                calculate_min_current_mh, axis=1
                            )

                            # Replace any remaining NaN or None values with 0
                            filtered_data = filtered_data.fillna(0)

                            # Re-arranging the columns as per the given order
                            final_columns = [
                                'spe', 'w4 rev', 'MINIMUM CURRENT MH', 
                                '1st on ms', 'CURRENT MH (1st on ms)', 'SUB1 (1st on ms)', 'CURRENT MH (SUB1 1st on ms)', 'SUB2 (1st on ms)', 'CURRENT MH (SUB2 1st on ms)',
                                '2nd on ms', 'CURRENT MH (2nd on ms)', 'SUB1 (2nd on ms)', 'CURRENT MH (SUB1 2nd on ms)', 'SUB2 (2nd on ms)', 'CURRENT MH (SUB2 2nd on ms)',
                                '3rd on ms', 'CURRENT MH (3rd on ms)', 'SUB1 (3rd on ms)', 'CURRENT MH (SUB1 3rd on ms)', 'SUB2 (3rd on ms)', 'CURRENT MH (SUB2 3rd on ms)',
                                '4th on ms', 'CURRENT MH (4th on ms)', 'SUB1 (4th on ms)', 'CURRENT MH (SUB1 4th on ms)', 'SUB2 (4th on ms)', 'CURRENT MH (SUB2 4th on ms)',
                                '5th on ms', 'CURRENT MH (5th on ms)', 'SUB1 (5th on ms)', 'CURRENT MH (SUB1 5th on ms)', 'SUB2 (5th on ms)', 'CURRENT MH (SUB2 5th on ms)',
                                'rev on ms', 'CURRENT MH (rev on ms)', 'SUB1 (rev on ms)', 'CURRENT MH (SUB1 rev on ms)', 'SUB2 (rev on ms)', 'CURRENT MH (SUB2 rev on ms)',
                                'cm on ls', 'CURRENT MH (cm on ls)', 'SUB1 (cm on ls)', 'CURRENT MH (SUB1 cm on ls)', 'SUB2 (cm on ls)', 'CURRENT MH (SUB2 cm on ls)',
                                'rev idler', 'CURRENT MH (rev idler)', 'SUB1 (rev idler)', 'CURRENT MH (SUB1 rev idler)', 'SUB2 (rev idler)', 'CURRENT MH (SUB2 rev idler)',
                                '3rd on ls', 'CURRENT MH (3rd on ls)', 'SUB1 (3rd on ls)', 'CURRENT MH (SUB1 3rd on ls)', 'SUB2 (3rd on ls)', 'CURRENT MH (SUB2 3rd on ls)',
                                '4th on ls', 'CURRENT MH (4th on ls)', 'SUB1 (4th on ls)', 'CURRENT MH (SUB1 4th on ls)', 'SUB2 (4th on ls)', 'CURRENT MH (SUB2 4th on ls)',
                                '5th on ls', 'CURRENT MH (5th on ls)', 'SUB1 (5th on ls)', 'CURRENT MH (SUB1 5th on ls)', 'SUB2 (5th on ls)', 'CURRENT MH (SUB2 5th on ls)',
                                'input shaft', 'CURRENT MH (input shaft)', 'SUB1 (input shaft)', 'CURRENT MH (SUB1 input shaft)', 'SUB2 (input shaft)', 'CURRENT MH (SUB2 input shaft)',
                                'main shaft', 'CURRENT MH (main shaft)', 'SUB1 (main shaft)', 'CURRENT MH (SUB1 main shaft)', 'SUB2 (main shaft)', 'CURRENT MH (SUB2 main shaft)',
                                'lay shaft', 'CURRENT MH (lay shaft)', 'SUB1 (lay shaft)', 'CURRENT MH (SUB1 lay shaft)', 'SUB2 (lay shaft)', 'CURRENT MH (SUB2 lay shaft)',
                                'hub 1/ 2', 'CURRENT MH (hub 1/ 2)', 'SUB1 (hub 1/ 2)', 'CURRENT MH (SUB1 hub 1/ 2)', 'SUB2 (hub 1/ 2)', 'CURRENT MH (SUB2 hub 1/ 2)',
                                'hub 3/4', 'CURRENT MH (hub 3/4)', 'SUB1 (hub 3/4)', 'CURRENT MH (SUB1 hub 3/4)', 'SUB2 (hub 3/4)', 'CURRENT MH (SUB2 hub 3/4)',
                                'hub 5/6', 'CURRENT MH (hub 5/6)', 'SUB1 (hub 5/6)', 'CURRENT MH (SUB1 hub 5/6)', 'SUB2 (hub 5/6)', 'CURRENT MH (SUB2 hub 5/6)',
                                'fdr', 'CURRENT MH (fdr)', 'SUB1 (fdr)', 'CURRENT MH (SUB1 fdr)', 'SUB2 (fdr)', 'CURRENT MH (SUB2 fdr)',
                                'sleeve 1/ 2', 'CURRENT MH (sleeve 1/ 2)', 'SUB1 (sleeve 1/ 2)', 'CURRENT MH (SUB1 sleeve 1/ 2)', 'SUB2 (sleeve 1/ 2)', 'CURRENT MH (SUB2 sleeve 1/ 2)',
                                'sleeve 3/4', 'CURRENT MH (sleeve 3/4)', 'SUB1 (sleeve 3/4)', 'CURRENT MH (SUB1 sleeve 3/4)', 'SUB2 (sleeve 3/4)', 'CURRENT MH (SUB2 sleeve 3/4)',
                                'sleeve 5/6', 'CURRENT MH (sleeve 5/6)', 'SUB1 (sleeve 5/6)', 'CURRENT MH (SUB1 sleeve 5/6)', 'SUB2 (sleeve 5/6)', 'CURRENT MH (SUB2 sleeve 5/6)',
                                'cone 1/2', 'CURRENT MH (cone 1/2)', 'SUB1 (cone 1/2)', 'CURRENT MH (SUB1 cone 1/2)', 'SUB2 (cone 1/2)', 'CURRENT MH (SUB2 cone 1/2)',
                                'cone 3/4', 'CURRENT MH (cone 3/4)', 'SUB1 (cone 3/4)', 'CURRENT MH (SUB1 cone 3/4)', 'SUB2 (cone 3/4)', 'CURRENT MH (SUB2 cone 3/4)',
                                'cone 5/6', 'CURRENT MH (cone 5/6)', 'SUB1 (cone 5/6)', 'CURRENT MH (SUB1 cone 5/6)', 'SUB2 (cone 5/6)', 'CURRENT MH (SUB2 cone 5/6)',
                                'cone 3', 'CURRENT MH (cone 3)', 'SUB1 (cone 3)', 'CURRENT MH (SUB1 cone 3)', 'SUB2 (cone 3)', 'CURRENT MH (SUB2 cone 3)',
                                'cone 4', 'CURRENT MH (cone 4)', 'SUB1 (cone 4)', 'CURRENT MH (SUB1 cone 4)', 'SUB2 (cone 4)', 'CURRENT MH (SUB2 cone 4)'
                            ]

                            # Reorder columns based on final_columns order
                            filtered_4_w_a = filtered_data[final_columns]

                            # Display the final DataFrame
                            st.dataframe(filtered_4_w_a)
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Processed Data"

                                # Write DataFrame to the worksheet
                            for row in dataframe_to_rows(filtered_4_w_a, index=False, header=True):
                                ws.append(row)

                                # Save the workbook to the BytesIO object
                            wb.save(output)
                            processed_file = output.getvalue()

                            st.download_button(
                                label="Download Processed Excel",
                                data=processed_file,
                                file_name="4_week_with_alternative.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            

                        else:
                            st.error('Error: Missing "P.NO", "SUB1", or "SUB2" columns in the Alternate Part Master sheet.')

        except Exception as e:
            st.error(f"An error occurred: {e}")


def four_week_without_alter():


# Title of the app
    st.title('4-week-without-alternative')

    # File uploader widget to upload the Excel file
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file is not None:
        # Load the Excel file
        try:
            # Load all sheet names and match case-insensitively by converting them to lowercase
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names

            # Convert sheet names to lowercase for case-insensitive comparison
            target_sheet_name_1 = "GB Requirement for Bal Month".lower()
            target_sheet_name_2 = "Part Raw Data".lower()
            target_sheet_name_3 = "Made Here Parts Calc".lower()

            matching_sheets_1 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_1]
            matching_sheets_2 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_2]
            matching_sheets_3 = [sheet for sheet in sheet_names if sheet.lower() == target_sheet_name_3]

            if matching_sheets_1 and matching_sheets_2 and matching_sheets_3:
                # Read the matched sheets
                df_1 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_1[0])
                df_2 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_2[0])
                df_3 = pd.read_excel(uploaded_file, sheet_name=matching_sheets_3[0])

                # Clean column names: remove leading/trailing spaces and make them case-insensitive
                df_1.columns = df_1.columns.str.strip().str.lower()
                df_2.columns = df_2.columns.str.strip().str.lower()
                df_3.columns = df_3.columns.str.strip().str.lower()

                # Check if 'spe' column exists in both sheets
                if 'spe' in df_1.columns and 'spe' in df_2.columns:
                    # Merge the two dataframes based on 'spe'
                    merged_data = pd.merge(df_1[['spe', 'w4 rev']], df_2, on='spe', how='inner')

                    # List of the required columns from the Part Raw Data sheet
                    required_columns = [
                        '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                        'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                        'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                        'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                    ]

                    # Check if all required columns are present
                    missing_columns = [col for col in required_columns if col not in merged_data.columns]
                    if missing_columns:
                        st.error(f"Missing columns in Part Raw Data: {', '.join(missing_columns)}")
                    else:
                        # Select only the relevant columns and add Serial Number starting from 1
                        filtered_data = merged_data[['spe', 'w4 rev'] + required_columns]

                        # Replace None or NaN with 0 in the filtered data
                        filtered_data = filtered_data.fillna(0)

                        # Add Serial Number starting from 1
                        filtered_data.insert(0, 'Serial Number', range(1, len(filtered_data) + 1))

                        # --- New functionality for "Made Here Parts Calc" sheet ---

                        # Extract P.NO and CURRENT MH columns from "Made Here Parts Calc"
                        if 'p.no' in df_3.columns and 'current mh' in df_3.columns:
                            made_here_parts_calc_df = df_3[['p.no', 'current mh']]

                            # Track remaining stock using P.NO
                            remaining_stock = made_here_parts_calc_df.set_index("p.no")["current mh"].to_dict()

                            # Function to calculate CURRENT MH and REMAINING MH row-wise for each part and column
                            def calculate_row_current_and_remaining(row, column):
                                part_id = row[column]
                                if part_id in remaining_stock:
                                    available_mh = remaining_stock[part_id]
                                    current_mh = min(row["w4 rev"], available_mh)  # Choose the min between W2 REV and available MH
                                    remaining_stock[part_id] -= current_mh  # Update remaining stock
                                    return current_mh
                                return 0

                            # Columns to process (the ones that start with 'F' like '1st on ms', '2nd on ms', etc.)
                            columns_to_process = [
                                '1st on ms', '2nd on ms', '3rd on ms', '4th on ms', '5th on ms', 'rev on ms',
                                'cm on ls', 'rev idler', '3rd on ls', '4th on ls', '5th on ls', 'input shaft',
                                'main shaft', 'lay shaft', 'hub 1/ 2', 'hub 3/4', 'hub 5/6', 'fdr', 'sleeve 1/ 2',
                                'sleeve 3/4', 'sleeve 5/6', 'cone 1/2', 'cone 3/4', 'cone 5/6', 'cone 3', 'cone 4'
                            ]

                            # Calculate CURRENT MH and REMAINING MH for all components
                            for col in columns_to_process:
                                # Calculate CURRENT MH for each part (using the `calculate_row_current_and_remaining` function)
                                filtered_data[f"CURRENT MH ({col})"] = filtered_data.apply(
                                    lambda row: calculate_row_current_and_remaining(row, col), axis=1
                                )

                                # Calculate REMAINING MH as W2 REV - CURRENT MH
                                filtered_data[f"REMAINING MH ({col})"] = (
                                    filtered_data["w4 rev"] - filtered_data[f"CURRENT MH ({col})"]
                                )

                            # --- Step 4: Calculate the minimum CURRENT MH for each row ---
                            def calculate_min_current_mh(row):
                                non_zero_values = [
                                    row[f"CURRENT MH ({col})"] for col in columns_to_process
                                    if row[f"CURRENT MH ({col})"] != 0
                                ]
                                return min(non_zero_values) if non_zero_values else 0

                            filtered_data["MINIMUM CURRENT MH"] = filtered_data.apply(
                                calculate_min_current_mh, axis=1
                            )

                            # Re-arranging the columns as per the given order
                            final_columns = [
                                'spe', 'w4 rev','MINIMUM CURRENT MH','1st on ms', 'CURRENT MH (1st on ms)', 'REMAINING MH (1st on ms)',
                                '2nd on ms', 'CURRENT MH (2nd on ms)', 'REMAINING MH (2nd on ms)', '3rd on ms', 'CURRENT MH (3rd on ms)', 'REMAINING MH (3rd on ms)',
                                '4th on ms', 'CURRENT MH (4th on ms)', 'REMAINING MH (4th on ms)', '5th on ms', 'CURRENT MH (5th on ms)', 'REMAINING MH (5th on ms)',
                                'rev on ms', 'CURRENT MH (rev on ms)', 'REMAINING MH (rev on ms)', 'cm on ls', 'CURRENT MH (cm on ls)', 'REMAINING MH (cm on ls)',
                                'rev idler', 'CURRENT MH (rev idler)', 'REMAINING MH (rev idler)', '3rd on ls', 'CURRENT MH (3rd on ls)', 'REMAINING MH (3rd on ls)',
                                '4th on ls', 'CURRENT MH (4th on ls)', 'REMAINING MH (4th on ls)', '5th on ls', 'CURRENT MH (5th on ls)', 'REMAINING MH (5th on ls)',
                                'input shaft', 'CURRENT MH (input shaft)', 'REMAINING MH (input shaft)', 'main shaft', 'CURRENT MH (main shaft)', 'REMAINING MH (main shaft)',
                                'lay shaft', 'CURRENT MH (lay shaft)', 'REMAINING MH (lay shaft)', 'hub 1/ 2', 'CURRENT MH (hub 1/ 2)', 'REMAINING MH (hub 1/ 2)',
                                'hub 3/4', 'CURRENT MH (hub 3/4)', 'REMAINING MH (hub 3/4)', 'hub 5/6', 'CURRENT MH (hub 5/6)', 'REMAINING MH (hub 5/6)',
                                'fdr', 'CURRENT MH (fdr)', 'REMAINING MH (fdr)', 'sleeve 1/ 2', 'CURRENT MH (sleeve 1/ 2)', 'REMAINING MH (sleeve 1/ 2)',
                                'sleeve 3/4', 'CURRENT MH (sleeve 3/4)', 'REMAINING MH (sleeve 3/4)', 'sleeve 5/6', 'CURRENT MH (sleeve 5/6)', 'REMAINING MH (sleeve 5/6)',
                                'cone 1/2', 'CURRENT MH (cone 1/2)', 'REMAINING MH (cone 1/2)', 'cone 3/4', 'CURRENT MH (cone 3/4)', 'REMAINING MH (cone 3/4)',
                                'cone 5/6', 'CURRENT MH (cone 5/6)', 'REMAINING MH (cone 5/6)', 'cone 3', 'CURRENT MH (cone 3)', 'REMAINING MH (cone 3)', 'cone 4', 'CURRENT MH (cone 4)', 'REMAINING MH (cone 4)'
                            ]

                            # Reorder the columns based on the final_columns list
                            filtered_4_wo_a= filtered_data[final_columns]

                            # Display the final mapped data with the CURRENT MH and REMAINING MH columns
                            st.write("Mapped Data from 'GB Requirement for Bal Month', 'Part Raw Data', and 'Made Here Parts Calc':")
                            st.dataframe(filtered_4_wo_a)
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Processed Data"

                                # Write DataFrame to the worksheet
                            for row in dataframe_to_rows(filtered_4_wo_a, index=False, header=True):
                                ws.append(row)

                                # Save the workbook to the BytesIO object
                            wb.save(output)
                            processed_file = output.getvalue()

                            st.download_button(
                                label="Download Processed Excel",
                                data=processed_file,
                                file_name="4_week_without_alternative.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                        else:
                            st.error('Columns "P.NO" or "CURRENT MH" not found in the "Made Here Parts Calc" sheet.')

                else:
                    st.error('Column "SPE" not found in one or both sheets.')

            else:
                if not matching_sheets_1:
                    st.error('Sheet "GB Requirement for Bal Month" not found in the provided Excel file.')
                if not matching_sheets_2:
                    st.error('Sheet "Part Raw Data" not found in the provided Excel file.')
                if not matching_sheets_3:
                    st.error('Sheet "Made Here Parts Calc" not found in the provided Excel file.')

        except Exception as e:
            st.error(f"Error reading the Excel file: {e}")


        

def Gbreq():
    st.title("GB Requirement For Bal Month")

    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file:
        try:
            workbook = pd.ExcelFile(uploaded_file)
            sheet_names = [name.lower() for name in workbook.sheet_names]

            if "monthly opening stock" not in sheet_names or "3 month plan" not in sheet_names:
                st.error("Ensure the Excel file has sheets named 'Monthly Opening Stock' and '3 Month Plan'.")
            else:
                os_sheet = pd.read_excel(workbook, sheet_name=[s for s in workbook.sheet_names if s.lower() == "monthly opening stock"][0])
                plan_sheet = pd.read_excel(workbook, sheet_name=[s for s in workbook.sheet_names if s.lower() == "3 month plan"][0])
                
                os_sheet.columns = os_sheet.columns.str.lower()
                plan_sheet.columns = plan_sheet.columns.str.lower()

                # Extract months from headers
                month_headers = list(
                    set(
                        [
                            col.split(" w")[0]
                            for col in plan_sheet.columns
                            if "w" in col
                        ]
                    )
                )

                selected_month = st.selectbox("Select Month", month_headers)

                if selected_month:
                    st.subheader(f"Results for {selected_month}")

                    results = []

                    for i, row in os_sheet.iterrows():
                        gb_value = row.get("gb", 0)
                        opening_stock = row.get("opening stock", 0)

                        w1_plan = plan_sheet.get(f"{selected_month} w1", pd.Series([0])).iloc[i] if i < len(plan_sheet) else 0
                        w2_plan = plan_sheet.get(f"{selected_month} w2", pd.Series([0])).iloc[i] if i < len(plan_sheet) else 0
                        w3_plan = plan_sheet.get(f"{selected_month} w3", pd.Series([0])).iloc[i] if i < len(plan_sheet) else 0
                        w4_plan = plan_sheet.get(f"{selected_month} w4", pd.Series([0])).iloc[i] if i < len(plan_sheet) else 0

                        w1_rev = max(0, w1_plan - opening_stock)
                        w1_excess = w1_rev - w1_plan

                        w2_rev_plan = w2_plan + w1_excess
                        w2_rev_plan_wn = max(0, w2_rev_plan)
                        week2_excess = w2_rev_plan - w2_plan

                        w3_rev_plan = w3_plan + week2_excess
                        w3_rev_plan_wn = max(0, w3_rev_plan)
                        week3_excess = w3_rev_plan - w3_plan

                        w4_rev_plan = w4_plan + week3_excess
                        w4_rev_plan_wn = max(0, w4_rev_plan)

                        results.append(
                            {
                                "GB": gb_value,
                                "Total": w1_plan + w2_plan + w3_plan + w4_plan,
                                "W1": w1_plan,
                                "W2": w2_plan,
                                "W3": w3_plan,
                                "W4": w4_plan,
                                "W1 Rev": w1_rev,
                                "W1 Excess": w1_excess,
                                "W2 Rev Plan": w2_rev_plan,
                                "W2 Rev Plan w/o Negative": w2_rev_plan_wn,
                                "Week 2 Excess / Less": week2_excess,
                                "W3 Rev Plan": w3_rev_plan,
                                "W3 Rev Plan w/o Negative": w3_rev_plan_wn,
                                "Week 3 Excess / Less": week3_excess,
                                "W4 Rev Plan": w4_rev_plan,
                                "W4 Rev Plan w/o Negative": w4_rev_plan_wn,
                            }
                        )

                    results_df_gb = pd.DataFrame(results)
                    results_df_gb.fillna(0, inplace=True)

                    st.dataframe(results_df_gb, use_container_width=True)

                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Processed Data"

                    # Write DataFrame to the worksheet
                    for row in dataframe_to_rows(results_df_gb, index=False, header=True):
                        ws.append(row)

                    # Save the workbook to the BytesIO object
                    wb.save(output)
                    processed_file = output.getvalue()

                    st.download_button(
                        label="Download Processed Excel",
                        data=processed_file,
                        file_name="GB Requirement For Bal Month.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        except Exception as e:
            st.error(f"Error processing file: {e}")


def Month():
    st.title("Monthly GB Requirement After OS")

    # File upload
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file:
        try:
            workbook = pd.ExcelFile(uploaded_file)
            
            # Convert sheet names to lowercase for case-insensitive comparison
            sheet_names_lower = {sheet.lower(): sheet for sheet in workbook.sheet_names}
            
            required_sheets = ["3 month plan", "monthly opening stock"]
            if not all(sheet in sheet_names_lower for sheet in required_sheets):
                st.error("Ensure the Excel file has sheets named 'Monthly Opening Stock' and '3 Month Plan' (case insensitive).")
            else:
                # Load sheets with case-insensitive sheet names
                plan_df = pd.read_excel(workbook, sheet_name=sheet_names_lower["3 month plan"])
                os_df = pd.read_excel(workbook, sheet_name=sheet_names_lower["monthly opening stock"])
                
                # Convert column names to lowercase for case-insensitive comparison
                plan_df.columns = map(str.lower, plan_df.columns)
                os_df.columns = map(str.lower, os_df.columns)
                
                # Extract unique months (case-insensitive column headers)
                month_headers = list(
                    {
                        header.split(" w")[0].strip()
                        for header in plan_df.columns
                        if " w" in header
                    }
                )
                
                selected_month = st.selectbox("Select Month", month_headers)
                
                if selected_month:
                    st.subheader(f"Results for {selected_month}")
                    processed_data = []
                    
                    for _, row in os_df.iterrows():
                        gb_value = row.get("gb", 0)
                        opening_stock = row.get("opening stock", 0)
                        remaining_stock = opening_stock
                        
                        row_result = {"GB": gb_value, "Opening Stock": opening_stock}
                        
                        for week in ["w1", "w2", "w3", "w4"]:
                            header = f"{selected_month} {week}"
                            if header in plan_df.columns:
                                week_plan = plan_df.loc[_, header] if _ < len(plan_df) else 0
                                fulfilled_plan = min(week_plan, remaining_stock)
                                unmet_plan = week_plan - fulfilled_plan
                                remaining_stock -= fulfilled_plan
                                
                                row_result[header] = week_plan
                                row_result[f"Plan for {week}"] = unmet_plan
                        
                        processed_data.append(row_result)
                        
                    # Rearrange columns
                    column_order = [
                        "GB", "Opening Stock",
                        f"{selected_month} w1", f"{selected_month} w2", f"{selected_month} w3", f"{selected_month} w4",
                        "Plan for w1", "Plan for w2", "Plan for w3", "Plan for w4"
                    ]
                    results_df = pd.DataFrame(processed_data)[column_order]
                    results_df.fillna(0, inplace=True)

                    st.write("### Calculated Results")
                    st.dataframe(results_df, use_container_width=True)

                    # Download button
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Processed Data"

                    # Write DataFrame to the worksheet
                    for row in dataframe_to_rows(results_df, index=False, header=True):
                        ws.append(row)

                    # Save the workbook to the BytesIO object
                    wb.save(output)
                    processed_file = output.getvalue()

                    st.download_button(
                        label="Download Processed Excel",
                        data=processed_file,
                        file_name="Month_GB_Requirement_After_OS.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        except Exception as e:
            st.error(f"An error occurred: {e}")
    
def map_wout_alt():
    st.title("Mapped set without alternative")

    # File uploader for user to upload an Excel file
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file:
        try:
            # Load the Excel file
            data = pd.ExcelFile(uploaded_file)

            # Convert sheet names to uppercase for case-insensitive matching
            sheet_names_upper = {sheet_name.upper(): sheet_name for sheet_name in data.sheet_names}

            # Required sheets with case-insensitive matching
            required_sheets = [
                "TODAY'S TENTATIVE PLAN", "PART RAW DATA", "NOMENCLATURE MASTER", "MADE HERE PARTS CALC"
            ]

            # Validate sheet names
            if not all(sheet in sheet_names_upper for sheet in required_sheets):
                missing_sheets = [sheet for sheet in required_sheets if sheet not in sheet_names_upper]
                st.error(f"Missing sheets: {', '.join(missing_sheets)}")
            else:
                # Load sheets
                tentative_plan_df = data.parse(sheet_names_upper["TODAY'S TENTATIVE PLAN"])
                nomenclature_master_df = data.parse(sheet_names_upper["NOMENCLATURE MASTER"])
                part_raw_data_df = data.parse(sheet_names_upper["PART RAW DATA"])
                made_here_parts_calc_df = data.parse(sheet_names_upper["MADE HERE PARTS CALC"])

                # Standardize column names to uppercase for case-insensitive matching
                tentative_plan_df.columns = tentative_plan_df.columns.str.strip().str.upper()
                nomenclature_master_df.columns = nomenclature_master_df.columns.str.strip().str.upper()
                part_raw_data_df.columns = part_raw_data_df.columns.str.strip().str.upper()
                made_here_parts_calc_df.columns = made_here_parts_calc_df.columns.str.strip().str.upper()

                # Ensure required columns exist
                required_columns = {
                    "TODAY'S TENTATIVE PLAN": ["MODEL", "QTY"],
                    "NOMENCLATURE MASTER": ["MODEL", "SPE"],
                    "PART RAW DATA": [
                        "SPE", "1ST ON MS", "2ND ON MS", "3RD ON MS", "4TH ON MS", "5TH ON MS", "REV ON MS", "CM ON LS",
                        "REV IDLER", "3RD ON LS", "4TH ON LS", "5TH ON LS", "INPUT SHAFT", "MAIN SHAFT", "LAY SHAFT",
                        "HUB 1/ 2", "HUB 3/4", "HUB 5/6", "FDR", "SLEEVE 1/ 2", "SLEEVE 3/4", "SLEEVE 5/6",
                        "CONE 1/2", "CONE 3/4", "CONE 5/6", "CONE 3", "CONE 4"
                    ],
                    "MADE HERE PARTS CALC": ["P.NO", "CURRENT MH"]
                }

                missing_columns = []
                for sheet_name, cols in required_columns.items():
                    df = {
                        "TODAY'S TENTATIVE PLAN": tentative_plan_df,
                        "NOMENCLATURE MASTER": nomenclature_master_df,
                        "PART RAW DATA": part_raw_data_df,
                        "MADE HERE PARTS CALC": made_here_parts_calc_df
                    }[sheet_name]

                    for col in cols:
                        if col not in df.columns:
                            missing_columns.append(f"{col} in {sheet_name}")

                if missing_columns:
                    st.error(f"Missing columns: {', '.join(missing_columns)}")
                else:
                    # Step 1: Map MODEL to SPE
                    tentative_plan_df = tentative_plan_df.merge(
                        nomenclature_master_df[["MODEL", "SPE"]], on="MODEL", how="left"
                    )

                    # Step 2: Map SPE to columns
                    columns_to_process = [
                        "1ST ON MS", "2ND ON MS", "3RD ON MS", "4TH ON MS", "5TH ON MS", "REV ON MS", "CM ON LS",
                        "REV IDLER", "3RD ON LS", "4TH ON LS", "5TH ON LS", "INPUT SHAFT", "MAIN SHAFT", "LAY SHAFT",
                        "HUB 1/ 2", "HUB 3/4", "HUB 5/6", "FDR", "SLEEVE 1/ 2", "SLEEVE 3/4", "SLEEVE 5/6",
                        "CONE 1/2", "CONE 3/4", "CONE 5/6", "CONE 3", "CONE 4"
                    ]

                    tentative_plan_df = tentative_plan_df.merge(
                        part_raw_data_df[["SPE"] + columns_to_process], on="SPE", how="left"
                    )

                    # Replace None or NaN values with 0
                    tentative_plan_df.fillna(0, inplace=True)

                    # Step 3: Track Remaining Stock
                    remaining_stock = made_here_parts_calc_df.set_index("P.NO")["CURRENT MH"].to_dict()

                    # Calculate CURRENT MH and REMAINING MH row-wise for each column
                    def calculate_row_current_and_remaining(row, column):
                        part_id = row[column]
                        if part_id in remaining_stock:
                            available_mh = remaining_stock[part_id]
                            current_mh = min(row["QTY"], available_mh)
                            remaining_stock[part_id] -= current_mh  # Update remaining stock
                            return current_mh
                        return 0

                    # Calculate CURRENT MH and REMAINING MH for all components
                    for col in columns_to_process:
                        tentative_plan_df[f"CURRENT MH ({col})"] = tentative_plan_df.apply(
                            lambda row: calculate_row_current_and_remaining(row, col), axis=1
                        )
                        # Calculate REMAINING MH as QTY - CURRENT MH
                        tentative_plan_df[f"REMAINING MH ({col})"] = (
                            tentative_plan_df["QTY"] - tentative_plan_df[f"CURRENT MH ({col})"]
                        )

                    # Step 4: Calculate the minimum CURRENT MH for each row
                    def calculate_min_current_mh(row):
                        non_zero_values = [
                            row[f"CURRENT MH ({col})"] for col in columns_to_process
                            if row[col] != 0
                        ]
                        return min(non_zero_values) if non_zero_values else 0

                    tentative_plan_df["MINIMUM CURRENT MH"] = tentative_plan_df.apply(
                        calculate_min_current_mh, axis=1
                    )

                    # Step 5: Select final columns for output
                    final_columns = ["MODEL", "SPE", "QTY", "MINIMUM CURRENT MH"]

                    for col in columns_to_process:
                        final_columns.extend([col, f"CURRENT MH ({col})", f"REMAINING MH ({col})"])

                    final_df = tentative_plan_df[final_columns]

                    # Add serial numbers starting from 1
                    final_df.reset_index(inplace=True, drop=True)
                    final_df.index = final_df.index + 1
                    final_df.index.name = "Serial Number"

                    # Step 6: Add a Total Row for MINIMUM CURRENT MH only
                    total_row = {col: "" for col in final_df.columns}  # Initialize with empty strings
                    total_row["MODEL"] = "TOTAL"  # Add label in the MODEL column
                    total_row["MINIMUM CURRENT MH"] = final_df["MINIMUM CURRENT MH"].sum()  # Sum for MINIMUM CURRENT MH

                    # Append the total row to the DataFrame
                    final_df_woal = pd.concat([final_df, pd.DataFrame([total_row])], ignore_index=True)

                    # Display the final DataFrame
                    st.write("### Processed Data (Detailed):")
                    st.dataframe(final_df_woal)

                    # Add download button for processed Excel
                    output = io.BytesIO()
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "Processed Data"

                    # Write DataFrame to Excel
                    for row in dataframe_to_rows(final_df_woal, index=True, header=True):
                        worksheet.append(row)

                    workbook.save(output)
                    output.seek(0)

                    st.download_button(
                        label="Download Processed Excel",
                        data=output,
                        file_name="processed_Without_Alternate_Part_Master.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        except Exception as e:
            st.error(f"An error occurred: {e}")
    else:
        st.info("Please upload an Excel file to get started.")


            
def Priority_Analysis_P_NO_with_WIP_Description_and_SUB1_Mapping():
    st.title("Priority Sheet")

    # File uploader
    data_file = st.file_uploader("Upload the Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if data_file is not None:
        try:
            # Read the uploaded Excel file
            excel_data = pd.ExcelFile(data_file)

            # Convert sheet names to uppercase for case-insensitive matching
            sheet_names_upper = {sheet_name.upper(): sheet_name for sheet_name in excel_data.sheet_names}

            def read_sheet(sheet_key):
                """Reads a sheet and makes column names uppercase for case-insensitive matching."""
                df = excel_data.parse(sheet_names_upper[sheet_key])
                df.columns = df.columns.str.strip().str.upper()
                return df

            # Check if 'Priority format' sheet exists
            part_no_column = None
            if 'PRIORITY FORMAT' in sheet_names_upper:
                priority_df = read_sheet('PRIORITY FORMAT')
                if 'P.NO' in priority_df.columns:
                    part_no_column = priority_df[['P.NO']].drop_duplicates().reset_index(drop=True)
                    part_no_column.index += 1  # Set serial numbers starting from 1
                    part_no_column.index.name = "Serial Number"
                else:
                    st.error("The 'P.NO' column was not found in the 'Priority format' sheet.")
            else:
                st.error("The 'Priority format' sheet was not found in the uploaded Excel file.")

            # Check if 'Made Here Parts Calc' sheet exists
            wip_data = None
            if 'MADE HERE PARTS CALC' in sheet_names_upper:
                made_here_df = read_sheet('MADE HERE PARTS CALC')
                required_columns = ['P.NO', 'HARD WIP', 'HT WIP', 'SOFT WIP', 'ROUGH WIP', 'WFT', 'DESC']
                missing_columns = [col for col in required_columns if col not in made_here_df.columns]

                if not missing_columns:
                    wip_data = made_here_df[required_columns].fillna(0)
                else:
                    st.error(f"Missing columns in 'Made Here Parts Calc': {', '.join(missing_columns)}")
            else:
                st.error("The 'Made Here Parts Calc' sheet was not found.")

            # Check if 'Alternate Part Master' sheet exists
            sub1_data = None
            if 'ALTERNATE PART MASTER' in sheet_names_upper:
                alternate_part_master_df = read_sheet('ALTERNATE PART MASTER')
                required_sub1_columns = ['P.NO', 'SUB1']
                missing_sub1_columns = [col for col in required_sub1_columns if col not in alternate_part_master_df.columns]

                if not missing_sub1_columns:
                    sub1_data = alternate_part_master_df[['P.NO', 'SUB1']].drop_duplicates().reset_index(drop=True)
                else:
                    st.error(f"Missing columns in 'Alternate Part Master': {', '.join(missing_sub1_columns)}")
            else:
                st.error("The 'Alternate Part Master' sheet was not found.")

            # Mapping Data
            if part_no_column is not None and wip_data is not None and sub1_data is not None:
                mapped_data = part_no_column.merge(wip_data, on='P.NO', how='left')
                mapped_data = mapped_data.merge(sub1_data, on='P.NO', how='left')
                mapped_data.fillna(0, inplace=True)

                sub1_wip_data = made_here_df[['P.NO', 'HARD WIP', 'HT WIP', 'SOFT WIP', 'ROUGH WIP', 'WFT']]
                sub1_wip_data.columns = ['SUB1', 'HARD WIP (2)', 'HT WIP (2)', 'SOFT WIP (2)', 'ROUGH WIP (2)', 'WFT (2)']
                mapped_data = mapped_data.merge(sub1_wip_data, on='SUB1', how='left').fillna(0)

                # Load Cycle Time Sheet if exists
                cycle_time_mapping = {}
                if 'CYCLE TIME SHEET' in sheet_names_upper:
                    cycle_time_df = read_sheet('CYCLE TIME SHEET')
                    if {'P.NO', 'CYCLE TIME'}.issubset(cycle_time_df.columns):
                        cycle_time_mapping = cycle_time_df.set_index('P.NO')['CYCLE TIME'].to_dict()

                # Calculate 1st Priority
                def calculate_1st_priority(row):
                    cycle_time = cycle_time_mapping.get(row['SUB1'], None)
                    if cycle_time is not None:
                        if cycle_time <= row['WFT']:
                            return "Hard-TG"
                    if row['WFT'] > 100:
                        return "Hard-TG"
                    return ""

                mapped_data['1st Priority'] = mapped_data.apply(calculate_1st_priority, axis=1)

                # Calculate 2nd Priority
                def calculate_2nd_priority(row):
                    if row['HT WIP'] > row['HARD WIP']:
                        return "Hard"
                    elif row['SOFT WIP'] + row['HT WIP'] > row['HARD WIP']:
                        return "HT"
                    elif row['SOFT WIP'] + row['HT WIP'] + row['ROUGH WIP'] + row['WFT'] > row['HARD WIP']:
                        return "Soft & HT"
                    elif row['SOFT WIP'] + row['HT WIP'] + row['ROUGH WIP'] + row['WFT'] < row['HARD WIP']:
                        return "Soft"
                    return "Rough"

                mapped_data['2nd Priority'] = mapped_data.apply(calculate_2nd_priority, axis=1)
                mapped_data['1st&2nd Priority'] = mapped_data['1st Priority'] + " & " + mapped_data['2nd Priority']

                st.subheader("Mapped Data: P.NO with WIP, Description, and SUB1 Columns")
                st.write(mapped_data)

                # Save to Excel
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Processed Data"
                for row in dataframe_to_rows(mapped_data, index=False, header=True):
                    ws.append(row)
                wb.save(output)
                processed_file = output.getvalue()

                st.download_button(
                    label="Download Priority sheet Excel",
                    data=processed_file,
                    file_name="Priority_sheet.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"An error occurred while processing the file: {e}")
    else:
        st.info("Please upload an Excel file to proceed.")


def process_part_matrix_master():
    st.title("Made Here Part Calculation")
    st.write("Upload an Excel file, and we'll process the 'Part Matrix Master', 'GB Requirement for Bal Month', and 'Date wise made here' sheets for you.")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file:
        try:
            excel_data = pd.ExcelFile(uploaded_file)

            # Convert sheet names to lowercase for case-insensitive matching
            available_sheets = {sheet.lower(): sheet for sheet in excel_data.sheet_names}
            required_sheets = ['part matrix master', 'gb requirement for bal month', 'date wise made here']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in available_sheets]

            if missing_sheets:
                st.error(f"Missing sheets in uploaded file: {', '.join(missing_sheets)}")
                return

            # Load dataframes using original sheet names from the uploaded file
            part_matrix_df = pd.read_excel(excel_data, sheet_name=available_sheets['part matrix master'])
            gb_requirement_df = pd.read_excel(excel_data, sheet_name=available_sheets['gb requirement for bal month'])
            date_wise_df = pd.read_excel(excel_data, sheet_name=available_sheets['date wise made here'])

            # Normalize column names to lowercase for case-insensitive processing
            part_matrix_df.columns = part_matrix_df.columns.str.lower()
            gb_requirement_df.columns = gb_requirement_df.columns.str.lower()
            date_wise_df.columns = date_wise_df.columns.str.lower()

            part_matrix_df.fillna(0, inplace=True)
            gb_requirement_df.fillna(0, inplace=True)

            # Handle 'Date wise made here' processing
            if 'date' not in date_wise_df.columns:
                st.error("'Date' column not found in 'Date wise made here' sheet.")
                return

            # Convert 'Date' to datetime and drop invalid rows
            date_wise_df['date'] = pd.to_datetime(date_wise_df['date'], errors='coerce')
            date_wise_df = date_wise_df.dropna(subset=['date'])

            # Fill NaN in other columns with 0, excluding 'Date'
            other_columns = [col for col in date_wise_df.columns if col != 'date']
            date_wise_df[other_columns] = date_wise_df[other_columns].fillna(0)

            unique_dates = date_wise_df['date'].drop_duplicates().sort_values()
            selected_date = st.selectbox("Select a Date", unique_dates)

            filtered_date_wise_df = date_wise_df[date_wise_df['date'] == selected_date]

            required_date_columns = {'current mh', 'hard wip', 'ht wip', 'soft wip', 'rough wip', 'hard wating for teeth'}
            missing_date_columns = required_date_columns - set(filtered_date_wise_df.columns)

            if not missing_date_columns:
                part_matrix_df['current mh'] = filtered_date_wise_df['current mh'].values
                part_matrix_df['hard wip'] = filtered_date_wise_df['hard wip'].values
                part_matrix_df['ht wip'] = filtered_date_wise_df['ht wip'].values
                part_matrix_df['soft wip'] = filtered_date_wise_df['soft wip'].values
                part_matrix_df['rough wip'] = filtered_date_wise_df['rough wip'].values
                part_matrix_df['hard wating for teeth'] = filtered_date_wise_df['hard wating for teeth'].values

                part_matrix_df.rename(columns={'hard wating for teeth': 'wft', 'current mh': 'store finished'}, inplace=True)
            else:
                st.warning(f"Missing columns in 'Date wise made here': {', '.join(missing_date_columns)}")
            
            # Convert object columns to numeric or datetime where possible
            for col in part_matrix_df.columns:
                if pd.api.types.is_object_dtype(part_matrix_df[col]):
                    try:
                        part_matrix_df[col] = pd.to_numeric(part_matrix_df[col], errors='ignore')
                    except:
                        pass

            for col in part_matrix_df.select_dtypes(include=['object']):
                try:
                    part_matrix_df[col] = pd.to_datetime(part_matrix_df[col], format='%d-%m-%Y', errors='ignore')
                    if part_matrix_df[col].dtype == 'object':
                        part_matrix_df[col] = pd.to_datetime(part_matrix_df[col], errors='ignore')
                except:
                    pass

            # Display processed data
            st.subheader("Processed 'Part Matrix Master' Sheet")
            st.dataframe(part_matrix_df)

            # Create downloadable Excel file
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed Data"

            for row in dataframe_to_rows(part_matrix_df, index=False, header=True):
                ws.append(row)

            wb.save(output)
            processed_file = output.getvalue()

            st.download_button(
                label="Download Processed Excel",
                data=processed_file,
                file_name="processed_part_matrix_master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"An error occurred: {e}")

def map_w_alt():
    st.title("mapped set with alternative")

    # File uploader for user to upload an Excel file
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx","xlsm","xlsb","xltx"])

    if uploaded_file:
        try:
            # Load the Excel file
            data = pd.ExcelFile(uploaded_file)

            # Normalize sheet names to lowercase for case-insensitive matching
            normalized_sheet_names = {sheet.lower(): sheet for sheet in data.sheet_names}

            # Required sheets
            required_sheets = [
                "today's tentative plan", "part raw data", "nomenclature master", "made here parts calc", "alternate part master"
            ]

            # Validate sheet names
            if not all(sheet in normalized_sheet_names for sheet in required_sheets):
                missing_sheets = [sheet for sheet in required_sheets if sheet not in normalized_sheet_names]
                st.error(f"Missing sheets: {', '.join(missing_sheets)}")
            else:
                # Load sheets
                tentative_plan_df = data.parse(normalized_sheet_names["today's tentative plan"])
                nomenclature_master_df = data.parse(normalized_sheet_names["nomenclature master"])
                part_raw_data_df = data.parse(normalized_sheet_names["part raw data"])
                made_here_parts_calc_df = data.parse(normalized_sheet_names["made here parts calc"])
                alternate_part_master_df = data.parse(normalized_sheet_names["alternate part master"])

                # Standardize column names
                tentative_plan_df.columns = tentative_plan_df.columns.str.strip().str.upper()
                nomenclature_master_df.columns = nomenclature_master_df.columns.str.strip().str.upper()
                part_raw_data_df.columns = part_raw_data_df.columns.str.strip().str.upper()
                made_here_parts_calc_df.columns = made_here_parts_calc_df.columns.str.strip().str.upper()
                alternate_part_master_df.columns = alternate_part_master_df.columns.str.strip().str.upper()

                # Ensure required columns exist
                required_columns = {
                    "Today's Tentative Plan": ["MODEL", "QTY"],
                    "Nomenclature Master": ["MODEL", "SPE"],
                    "Part Raw Data": [
                        "SPE", "1ST ON MS", "2ND ON MS", "3RD ON MS", "4TH ON MS", "5TH ON MS", "REV ON MS", "CM ON LS",
                        "REV IDLER", "3RD ON LS", "4TH ON LS", "5TH ON LS", "INPUT SHAFT", "MAIN SHAFT", "LAY SHAFT",
                        "HUB 1/ 2", "HUB 3/4", "HUB 5/6", "FDR", "SLEEVE 1/ 2", "SLEEVE 3/4", "SLEEVE 5/6",
                        "CONE 1/2", "CONE 3/4", "CONE 5/6", "CONE 3", "CONE 4"
                    ],
                    "Made Here Parts Calc": ["P.NO", "CURRENT MH"],
                    "Alternate Part Master": ["P.NO", "SUB1", "SUB2"]
                }

                missing_columns = []
                for sheet_name, cols in required_columns.items():
                    df = {
                        "Today's Tentative Plan": tentative_plan_df,
                        "Nomenclature Master": nomenclature_master_df,
                        "Part Raw Data": part_raw_data_df,
                        "Made Here Parts Calc": made_here_parts_calc_df,
                        "Alternate Part Master": alternate_part_master_df
                    }[sheet_name]

                    for col in cols:
                        if col.upper() not in df.columns:
                            missing_columns.append(f"{col} in {sheet_name}")

                if missing_columns:
                    st.error(f"Missing columns: {', '.join(missing_columns)}")
                else:
                    # Step 1: Map MODEL to SPE
                    tentative_plan_df = tentative_plan_df.merge(
                        nomenclature_master_df[["MODEL", "SPE"]], on="MODEL", how="left"
                    )

                    # Step 2: Map SPE to columns
                    columns_to_process = [
                        "1ST ON MS", "2ND ON MS", "3RD ON MS", "4TH ON MS", "5TH ON MS", "REV ON MS", "CM ON LS",
                        "REV IDLER", "3RD ON LS", "4TH ON LS", "5TH ON LS", "INPUT SHAFT", "MAIN SHAFT", "LAY SHAFT",
                        "HUB 1/ 2", "HUB 3/4", "HUB 5/6", "FDR", "SLEEVE 1/ 2", "SLEEVE 3/4", "SLEEVE 5/6",
                        "CONE 1/2", "CONE 3/4", "CONE 5/6", "CONE 3", "CONE 4"
                    ]

                    tentative_plan_df = tentative_plan_df.merge(
                        part_raw_data_df[["SPE"] + columns_to_process], on="SPE", how="left"
                    )

                    # Replace None or NaN values with 0
                    tentative_plan_df.fillna(0, inplace=True)

                    # Step 3: Map P.NO to SUB1 and SUB2 for each column
                    alternate_part_dict = alternate_part_master_df.set_index("P.NO")[["SUB1", "SUB2"]].to_dict("index")

                    for col in columns_to_process:
                        tentative_plan_df[f"SUB1 ({col})"] = tentative_plan_df[col].map(
                            lambda x: alternate_part_dict[x]["SUB1"] if x in alternate_part_dict else 0
                        )
                        tentative_plan_df[f"SUB2 ({col})"] = tentative_plan_df[col].map(
                            lambda x: alternate_part_dict[x]["SUB2"] if x in alternate_part_dict else 0
                        )

                    # Step 4: Calculate CURRENT MH row-wise for each column
                    remaining_stock = made_here_parts_calc_df.set_index("P.NO")["CURRENT MH"].to_dict()

                    def calculate_row_current(row, column):
                        key = row[column]
                        if key in remaining_stock:
                            available_mh = remaining_stock[key]
                            used_mh = min(row["QTY"], available_mh)
                            remaining_stock[key] -= used_mh
                            return used_mh
                        return 0

                    for col in columns_to_process:
                        tentative_plan_df[f"CURRENT MH ({col})"] = tentative_plan_df.apply(
                            lambda row: calculate_row_current(row, col), axis=1
                        )

                    # Step 5: Calculate CURRENT MH for SUB1
                    for col in columns_to_process:
                        tentative_plan_df[f"CURRENT MH (SUB1 {col})"] = tentative_plan_df.apply(
                            lambda row: max(row["QTY"] - row[f"CURRENT MH ({col})"], 0), axis=1
                        )

                    # Step 6: Calculate CURRENT MH for SUB2
                    for col in columns_to_process:
                        tentative_plan_df[f"CURRENT MH (SUB2 {col})"] = tentative_plan_df.apply(
                            lambda row: max(row[f"CURRENT MH ({col})"] + row[f"CURRENT MH (SUB1 {col})"] - row["QTY"], 0), axis=1
                        )

                    # Step 7: Calculate the minimum CURRENT MH for each row, excluding zero-value columns
                    def calculate_min_current_mh(row):
                        non_zero_values = [
                            row[f"CURRENT MH ({col})"] for col in columns_to_process
                            if row[col] != 0
                        ]
                        return min(non_zero_values) if non_zero_values else 0

                    tentative_plan_df["MINIMUM CURRENT MH"] = tentative_plan_df.apply(
                        calculate_min_current_mh, axis=1
                    )

                    # Step 8: Select final columns for output
                    final_columns = [
                        "MODEL", "SPE", "QTY", "MINIMUM CURRENT MH"
                    ]

                    for col in columns_to_process:
                        final_columns.extend([
                            col, f"CURRENT MH ({col})", f"SUB1 ({col})", f"CURRENT MH (SUB1 {col})", f"SUB2 ({col})", f"CURRENT MH (SUB2 {col})"
                        ])

                    final_df = tentative_plan_df[final_columns]

                    # Replace None or NaN values in the final DataFrame with 0
                    final_df.fillna(0, inplace=True)

                    # Add serial numbers starting from 1
                    final_df.reset_index(inplace=True, drop=True)
                    final_df.index = final_df.index + 1
                    final_df.index.name = "Serial Number"

                    # Step 9: Add a Total Row for MINIMUM CURRENT MH only
                    total_row = {col: 0 for col in final_df.columns}  # Initialize with zeros
                    total_row["MODEL"] = "TOTAL"  # Add label in the MODEL column
                    total_row["MINIMUM CURRENT MH"] = final_df["MINIMUM CURRENT MH"].sum()  # Sum for MINIMUM CURRENT MH

                    # Append the total row to the DataFrame
                    final_df_wal = pd.concat([final_df, pd.DataFrame([total_row])], ignore_index=True)

                    # Display the final DataFrame
                    st.write("### Processed Data (Detailed):")
                    st.dataframe(final_df_wal)

                    # Provide a download button for the processed data
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        final_df_wal.to_excel(writer, index=True, sheet_name="Processed Data")
                    st.download_button(
                        label="Download Processed Data",
                        data=output.getvalue(),
                        file_name="Processed_With_Alternate_Part_Master.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        except Exception as e:
            st.error(f"An error occurred: {e}")
    else:
        st.info("Please upload an Excel file to get started.")

        


def main():
    st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: #ffffff;
    }
    
    [data-testid="stSidebar"] {
        background: linear-gradient(195deg, #1a1a2e 0%, #16213e 100%) !important;
        border-right: 1px solid #2a529850;
    }

    /* Text color adjustments */
    .stTextInput>label, .stNumberInput>label, .stSelectbox>label,
    .stRadio>label, .stMarkdown, .stTitle {
        color: #ffffff !important;
    }

    /* Input field hover and focus effects */
    .stTextInput input, .stNumberInput input, .stTextArea textarea {
        transition: all 0.3s ease !important;
        background-color: rgba(125, 214, 235, 0.57) !important;
        border: 1px solidrgba(200, 12, 217, 0.31) !important;
    }

    .stTextInput input:hover, .stNumberInput input:hover, .stTextArea textarea:hover {
        background-color: rgba(255, 255, 255, 0.94) !important;
        transform: scale(1.02);
    }

    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        color:black;
        
       
        transform: scale(1.02);
    }

    /* Button styling */
.stButton>button {
    background: linear-gradient(45deg, #4CAF50 0%, #45a049 100%);
    color: white !important;
    border: none;
    border-radius: 5px;
    padding: 10px 24px;
    transition: all 0.3s ease !important;
}

.stButton>button:hover {
    background: linear-gradient(45deg, #2196F3 0%, #1976D2 100%) !important;
    transform: scale(1.05);
    opacity: 0.9;
}

    /* Container styling */
    .stContainer {
        background-color: rgba(255, 255, 255, 0.1) !important;
        border-radius: 10px;
        padding: 20px;
        margin: 10px 0;
        backdrop-filter: blur(5px);
    }

    /* Dataframe styling */
    .dataframe {
        background-color: rgba(23, 137, 195, 0.1) !important;
    }

    /* Hover effects */
    .stButton>button:hover {
        transform: scale(1.05);
        opacity: 0.9;
    }
    </style>
    """,
    unsafe_allow_html=True
    )
    
    
    st.sidebar.title("Navigation")
    menu = ["Login", "Register", "Logout", "Dashboard"]
    choice = st.sidebar.selectbox("Menu", menu)

    if choice == "Login":
        login()
    elif choice == "Register":
        register()
    elif choice == "Logout":
        logout()
    elif choice == "Dashboard":
        app_functionality()

# Define the other functions here (e.g., Gbreq, Month, etc.)

if __name__ == "__main__":
    main()
